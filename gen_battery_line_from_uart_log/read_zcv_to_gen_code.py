#coding=utf-8
import os
import pickle
import codecs
import sys
import xlrd
from openpyxl import load_workbook
import xlsxwriter
import re
import operator
import random
import cairo
import pycha.scatter
import pycha.bar
import jlink_gen_code as jlinkgc

time_soc_dataset=[];
time_count_name=[];
time_count_valset=[];
time_ocv_dataset=[];
soc_r_dataset=[];

soc_for_sprd=1;
nums_for_mtk=77;
#define for target


out_tuple_num = 100;
last_ocv = 3500;#to vbat is 3400.
target_base_time = 0;

def load_xlsx_data(f_path, sheet_name, soc_col, vol_col, r_col, row_s, row_e, in_r, dis_c):
    print f_path
    rb = load_workbook(f_path,data_only=True);
    print rb.get_sheet_names();
    
    for cur_sheet in rb.get_sheet_names():
        #print cur_sheet
        if cur_sheet == sheet_name:
            print rb[cur_sheet]
            rs = rb[cur_sheet]
            target_base_time = rs.cell(row=row_s, column=soc_col, value=10).value
            for rowidx in range(row_s, row_e):
                soc = rs.cell(row=rowidx, column=soc_col, value=10).value
                vol = rs.cell(row=rowidx, column=vol_col, value=10).value
                ocvr = rs.cell(row=rowidx, column=r_col, value=10).value
                print vol, ocvr, soc
                tmp_ocv = int(vol);

                if tmp_ocv > 3300:
                    tmpval = (soc,int(tmp_ocv-3300));
                    time_ocv_dataset.append(tmpval);
                    tmpval = (soc, ocvr);
                    soc_r_dataset.append(tmpval);


# gen png
def scatterplotChart(output, in_dataset):

    surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, 1920, 1080)
    #surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, 6920, 4080)

    top = 2
    dataSet = (
        ('points 1', [(i, random.random() * float(top)) for i in range(top)]),
        )
    #print dataSet
    dataSet=(('points 1',in_dataset),);
    #print dataSet
    options = {
        'background': {
            'color': '#eeeeff',
            'lineColor': '#444444',
        },
        'colorScheme': {
            'name': 'rainbow',
            'args': {
                'initialColor': 'blue',
            },
        },
        'legend': {
            'hide': True,
        },
        'title': u'Scatter plot',
    }
    chart = pycha.scatter.ScatterplotChart(surface, options)

    chart.addDataset(dataSet)
    chart.render()

    surface.write_to_png(output)


def read_xls_to_gen_ocv_table_same_r(in_f, sheet_name, soc_col, vol_col, r_col, row_s, row_e, in_r, dis_c):
    target_sheet_name=sheet_name
    target_vol_col=vol_col
    target_soc_col=soc_col
    target_row_start=row_s
    target_row_end=row_e
    dis_c=dis_c;
    xls_i_r = in_r

    print sys.argv
    print sys.path[0]
    curpath=sys.path[0]
    f_path=in_f
    print f_path
    input_dir=in_f[:in_f.rfind('/')]
    print input_dir

    load_xlsx_data(f_path, sheet_name, soc_col, vol_col, r_col, row_s, row_e, in_r, dis_c)

    print time_ocv_dataset;
    print soc_r_dataset;
#    scatterplotChart('xls_dis_curve.png', time_soc_dataset);
    scatterplotChart('read_zcv_soc_dis_curve.png', time_ocv_dataset);

####################################   print soc ###################################
    cnt=0;
    mtk_out=""
    cur_cnt=0;
    for vals in time_ocv_dataset:
        if vals[0] >= cnt:
            mtk_out= "%s{%d, %d},\n" % (mtk_out, cnt, vals[1]+3300)#, vals[1]+3300-in_r*dis_c/1000)
            cnt=cnt+2;
            cur_cnt=cur_cnt+1;
    #print "{%d, %d}," % (cnt, vals[1]+3300)

    print cur_cnt;
    for i in range(cur_cnt, nums_for_mtk):
        mtk_out= "%s{%d, %d},\n" % (mtk_out, 100, 3100)


####################################   print r ###################################
    mtk_out= "%sstart r table\n" % mtk_out

    cnt=0;
    cur_cnt=0;
    for idx in range(len(soc_r_dataset)):
        if soc_r_dataset[idx][0] >= cnt:
            mtk_out="%s{%d, %d},\n" % (mtk_out, soc_r_dataset[idx][1], time_ocv_dataset[idx][1]+3300)#, vals[1]+3300-in_r*dis_c/1000)
            cnt=cnt+2;
            cur_cnt=cur_cnt+1;

    for i in range(cur_cnt, nums_for_mtk):
        mtk_out= "%s{%d, %d},\n" % (mtk_out, soc_r_dataset[len(soc_r_dataset)-2][1], 3100)

    out_file=input_dir + "/read_ocv_out.h"
    genf=open(out_file, 'w+');
    genf.write(mtk_out);
    genf.close();

####################################   print for sprd ###################################
    if soc_for_sprd == 1:
        cnt=0
        outs=""
        for vals in time_ocv_dataset:
            if vals[0] >= cnt:
                outs = "%s %d " % (outs, 100-cnt)#, vals[1]+3300-in_r*dis_c/1000)
                cnt=cnt+2;
        cnt=0
        outs =  "%s\n" % (outs)#, vals[1]+3300-in_r*dis_c/1000)
        for vals in time_ocv_dataset:
            if vals[0] >= cnt:
                outs = "%s %d " % (outs, vals[1]+3300)#, vals[1]+3300-in_r*dis_c/1000)
                cnt=cnt+2;
        print outs
        print len(time_ocv_dataset)


#read_xls_to_gen_ocv_table_same_r("/home/manjusaka/work_data/b960/battery/battery/zcv.xlsx", "Sheet3", 8, 3, 9, 3, 74, 75, 2000);
#read_xls_to_gen_ocv_table_same_r("/home/manjusaka/work_data/K960/battery/lingyun_3100/tt.xlsx", "Sheet1", 9, 6, 1597, 3226, 60, 1400);

#read_xls_to_gen_ocv_table_same_r("/home/manjusaka/work_data/S960/battery/plr_d1_jjy/zcv.xlsx", "shee", 13, 9, 12, 2, 109, 60, 2000);
#read_xls_to_gen_ocv_table_same_r("/home/manjusaka/work_data/S525/battery/bk.xlsx", "ZCV", 6, 2, 7, 2, 103, 60, 2000);
read_xls_to_gen_ocv_table_same_r("/home/manjusaka/work_data/S505/BAT/TCL/zzcv.xlsx", "ZCV", 6, 2, 7, 3, 103, 66, 2000);
