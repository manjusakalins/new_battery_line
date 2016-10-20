#coding=utf-8
import os
import pickle
import codecs
import sys
import xlrd
from openpyxl import load_workbook
import xlsxwriter
import re
import operator
import random
import cairo
import pycha.scatter
import pycha.bar
import jlink_gen_code as jlinkgc

time_soc_dataset=[];
time_count_name=[];
time_count_valset=[];
time_ocv_dataset=[];
soc_r_dataset=[];

soc_for_sprd=1;
nums_for_mtk=77;
#define for target


out_tuple_num = 100;
last_ocv = 3500;#to vbat is 3400.
target_base_time = 0;

def load_xlsx_data(f_path, sheet_name, soc_col, vol_col, r_col, row_s, row_e, in_r, dis_c):
    out_data_set=[];
    rb = load_workbook(f_path,data_only=True);
    print rb.get_sheet_names();
    
    for cur_sheet in rb.get_sheet_names():
        #print cur_sheet
        if cur_sheet == sheet_name:
            print rb[cur_sheet]
            rs = rb[cur_sheet]
            target_base_time = rs.cell(row=row_s, column=soc_col, value=10).value
            for rowidx in range(row_s, row_e):
                soc = rs.cell(row=rowidx, column=soc_col, value=10).value
                vol = rs.cell(row=rowidx, column=vol_col, value=10).value
                ocvr = rs.cell(row=rowidx, column=r_col, value=10).value
                print vol, ocvr, soc

                ###### FIXME: maybe we need change cell data value here with differce xlsx #####
                tmp_ocv = int(vol);
                tmp_r = int(ocvr);
                ################################################################################

                if tmp_ocv > 3300:
                    tmpval = (soc,int(tmp_ocv-3300));
                    time_ocv_dataset.append(tmpval);
                    tmpval = (soc, ocvr);
                    soc_r_dataset.append(tmpval);
                if tmp_ocv > 3200:
                    tmp_dict={}
                    tmp_dict["ocv"] = tmp_ocv
                    tmp_dict["r"] = tmp_r
                    tmp_dict["soc"] = soc
                    out_data_set.append(tmp_dict);
                    
    return out_data_set;
# gen png
def scatterplotChart(output, in_dataset):

    surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, 1920, 1080)
    #surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, 6920, 4080)

    top = 2
    dataSet = (
        ('points 1', [(i, random.random() * float(top)) for i in range(top)]),
        )
    #print dataSet
    dataSet=(('points 1',in_dataset),);
    #print dataSet
    options = {
        'background': {
            'color': '#eeeeff',
            'lineColor': '#444444',
        },
        'colorScheme': {
            'name': 'rainbow',
            'args': {
                'initialColor': 'blue',
            },
        },
        'legend': {
            'hide': True,
        },
        'title': u'Scatter plot',
    }
    chart = pycha.scatter.ScatterplotChart(surface, options)

    chart.addDataset(dataSet)
    chart.render()

    surface.write_to_png(output)


def read_xls_to_gen_ocv_table_same_r(in_f, sheet_name, soc_col, vol_col, r_col, row_s, row_e, in_r, dis_c):

    curpath=sys.path[0]
    f_path=in_f
    input_dir=in_f[:in_f.rfind('/')]
    print input_dir
    load_xlsx_data(f_path, sheet_name, soc_col, vol_col, r_col, row_s, row_e, in_r, dis_c)

    #print time_ocv_dataset;
    #print soc_r_dataset;
    scatterplotChart('read_zcv_soc_dis_curve.png', time_ocv_dataset);

    sets=load_xlsx_data(f_path, sheet_name, soc_col, vol_col, r_col, row_s, row_e, in_r, dis_c)
    print sets
    jlinkgc.write_out_mtk_header(input_dir,"ydsm", "tcl", 2100, sets)
    jlinkgc.write_out_mtk_dtsi(input_dir,"ydsm", "tcl", 2100, sets)
    jlinkgc.write_out_sprd_dtsi(input_dir,"ydsm", "tcl", 2100, sets)
    

read_xls_to_gen_ocv_table_same_r("/home/manjusaka/work_data/S525/battery/bk.xlsx", "ZCV", 6, 2, 7, 2, 105, 60, 2000);
