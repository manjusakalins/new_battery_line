#coding=utf-8
import os
import pickle
import codecs
import sys
import xlrd
from openpyxl import load_workbook
import xlsxwriter
import re
import operator
import random
import cairo
import pycha.scatter
import pycha.bar


time_soc_dataset=[];
time_count_name=[];
time_count_valset=[];
time_ocv_dataset=[];

#define for target


out_tuple_num = 100;
last_ocv = 3500;#to vbat is 3400.
target_base_time = 0;

def load_xlsx_data(f_path, sheet_name, time_col, vol_col, row_s, row_e, in_r, dis_c, vol_base):
    print f_path
    rb = load_workbook(f_path);
    print rb.get_sheet_names();
    
    for cur_sheet in rb.get_sheet_names():
#        print cur_sheet
        if cur_sheet == sheet_name:
            print rb[cur_sheet]
            rs = rb[cur_sheet]
            target_base_time = rs.cell(row=row_s, column=time_col, value=10).value
            for rowidx in range(row_s, row_e):
                time = rs.cell(row=rowidx, column=time_col, value=10).value
                vol = rs.cell(row=rowidx, column=vol_col, value=10).value * vol_base
#                print d
                tmpval = (time - target_base_time,int(vol-2700));
                time_soc_dataset.append(tmpval);
                tmp_ocv = vol+dis_c*in_r/1000
                if tmp_ocv > 3485:
                    tmpval = (time - target_base_time,int(tmp_ocv-3300));
                    time_ocv_dataset.append(tmpval);


# gen png
def scatterplotChart(output, in_dataset):

    surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, 1920, 1080)
    #surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, 6920, 4080)

    top = 2
    dataSet = (
        ('points 1', [(i, random.random() * float(top)) for i in range(top)]),
        )
    #print dataSet
    dataSet=(('points 1',in_dataset),);
    #print dataSet
    options = {
        'background': {
            'color': '#eeeeff',
            'lineColor': '#444444',
        },
        'colorScheme': {
            'name': 'rainbow',
            'args': {
                'initialColor': 'blue',
            },
        },
        'legend': {
            'hide': True,
        },
        'title': u'Scatter plot',
    }
    chart = pycha.scatter.ScatterplotChart(surface, options)

    chart.addDataset(dataSet)
    chart.render()

    surface.write_to_png(output)


def read_xls_to_gen_ocv_table_same_r(in_f, sheet_name, time_col, vol_col, row_s, row_e, in_r, dis_c, vol_base):
    target_sheet_name=sheet_name
    target_vol_col=vol_col
    target_time_col=time_col
    target_row_start=row_s
    target_row_end=row_e
    dis_c=dis_c;
    xls_i_r = in_r

    print sys.argv
    print sys.path[0]
    curpath=sys.path[0]
    f_path=in_f
    print f_path
    
    load_xlsx_data(f_path, sheet_name, time_col, vol_col, row_s, row_e, in_r, dis_c, vol_base);
    print time_soc_dataset;
    print time_ocv_dataset;
    scatterplotChart('xls_dis_curve.png', time_soc_dataset);
    scatterplotChart('xls_dis_curve_same_r.png', time_ocv_dataset);
    
    total_time = time_ocv_dataset[len(time_ocv_dataset)-1][0]
    print total_time
    per_unit = float(total_time)/float(out_tuple_num);
    print per_unit
    cnt=0;
    for vals in time_ocv_dataset:
        if vals[0] >= cnt*per_unit:
            print "{%d, %d}," % (cnt, vals[1]+3300)#, vals[1]+3300-in_r*dis_c/1000)
            cnt=cnt+2;
    print "{%d, %d}," % (cnt, time_ocv_dataset[len(time_ocv_dataset)-1][1]+3300)
    #print r
    cnt=0
    for vals in time_ocv_dataset:
        if vals[0] >= cnt*per_unit:
            print "{%d, %d}," % (in_r, vals[1]+3300)#, vals[1]+3300-in_r*dis_c/1000)
            cnt=cnt+2;
    
#read_xls_to_gen_ocv_table_same_r("/home/manjusaka/work_data/I960/battery/i109_battery_4000/i109.xlsx", "Record", 3, 5, 1140, 2730, 75, 2000,1);
#read_xls_to_gen_ocv_table_same_r("/home/manjusaka/work_data/K960/battery/lingyun_3100/tt.xlsx", "Sheet1", 9, 6, 1597, 3226, 60, 1400,1);
read_xls_to_gen_ocv_table_same_r("/home/manjusaka/work_data/K960/battery/yinongcheng_4200/2line.xlsx", "TestData", 1, 5, 592, 853, 80, 2000, 1000);
